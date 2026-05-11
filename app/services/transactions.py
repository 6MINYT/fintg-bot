from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from sqlalchemy import Select, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.categories import category_label
from app.core.types import TransactionType
from app.db.models import BudgetLimit, Transaction, User
from app.services.parser import ParsedTransaction


async def get_or_create_user(
    session: AsyncSession,
    telegram_id: int,
    username: str | None,
    first_name: str | None,
) -> User:
    user = await session.scalar(select(User).where(User.telegram_id == telegram_id))
    if user:
        user.username = username
        user.first_name = first_name
        if not user.default_currency:
            user.default_currency = "PLN"
        return user

    user = User(telegram_id=telegram_id, username=username, first_name=first_name)
    session.add(user)
    await session.flush()
    return user


async def list_users(session: AsyncSession) -> list[User]:
    rows = await session.scalars(select(User).order_by(User.id))
    return list(rows)


async def user_activity_stats(
    session: AsyncSession,
    now: datetime,
) -> list[tuple[User, int, int, int, int, datetime | None, date | None]]:
    today_from = datetime.combine(now.date(), datetime.min.time(), tzinfo=now.tzinfo)
    active_7_from = now - timedelta(days=7)
    active_30_from = now - timedelta(days=30)
    last_created_at = func.max(Transaction.created_at)
    rows = await session.execute(
        select(
            User,
            func.count(Transaction.id),
            func.count(Transaction.id).filter(Transaction.created_at >= today_from),
            func.count(Transaction.id).filter(Transaction.created_at >= active_7_from),
            func.count(Transaction.id).filter(Transaction.created_at >= active_30_from),
            last_created_at,
            func.max(Transaction.occurred_on),
        )
        .outerjoin(Transaction, Transaction.user_id == User.id)
        .group_by(User.id)
        .order_by(last_created_at.desc().nullslast(), User.id)
    )
    return [
        (user, int(total or 0), int(today or 0), int(last_7 or 0), int(last_30 or 0), last_tx_created_at, last_tx_date)
        for user, total, today, last_7, last_30, last_tx_created_at, last_tx_date in rows.all()
    ]


async def update_user_default_currency(session: AsyncSession, user: User, currency: str) -> User:
    user.default_currency = currency
    await session.flush()
    return user


async def add_transaction(
    session: AsyncSession,
    user: User,
    parsed: ParsedTransaction,
    raw_text: str,
) -> Transaction:
    transaction = Transaction(
        user_id=user.id,
        user_tx_number=await _next_user_tx_number(session, user),
        type=parsed.type,
        amount=parsed.amount,
        currency=parsed.currency,
        category=parsed.category,
        merchant=parsed.merchant,
        note=parsed.note,
        occurred_on=parsed.occurred_on,
        raw_text=raw_text,
    )
    session.add(transaction)
    await session.flush()
    return transaction


async def _next_user_tx_number(session: AsyncSession, user: User) -> int:
    max_number = await session.scalar(
        select(func.coalesce(func.max(Transaction.user_tx_number), 0)).where(Transaction.user_id == user.id)
    )
    return int(max_number or 0) + 1


async def get_last_transaction(session: AsyncSession, user: User) -> Transaction | None:
    return await session.scalar(
        select(Transaction)
        .where(Transaction.user_id == user.id)
        .order_by(Transaction.created_at.desc(), Transaction.id.desc())
        .limit(1)
    )


async def get_transaction_by_id(session: AsyncSession, user: User, transaction_id: int) -> Transaction | None:
    return await session.scalar(
        select(Transaction).where(
            Transaction.user_tx_number == transaction_id,
            Transaction.user_id == user.id,
        )
    )


async def list_recent_transactions(session: AsyncSession, user: User, limit: int = 10) -> list[Transaction]:
    rows = await session.scalars(
        select(Transaction)
        .where(Transaction.user_id == user.id)
        .order_by(Transaction.user_tx_number.desc(), Transaction.created_at.desc(), Transaction.id.desc())
        .limit(limit)
    )
    return list(rows)


async def list_transactions_since(
    session: AsyncSession,
    user: User,
    from_date: date,
    limit: int = 50,
) -> list[Transaction]:
    rows = await session.scalars(
        select(Transaction)
        .where(
            Transaction.user_id == user.id,
            Transaction.occurred_on >= from_date,
        )
        .order_by(Transaction.occurred_on.desc(), Transaction.user_tx_number.desc(), Transaction.id.desc())
        .limit(limit)
    )
    return list(rows)


async def update_transaction(
    session: AsyncSession,
    transaction: Transaction,
    parsed: ParsedTransaction,
    raw_text: str,
) -> Transaction:
    transaction.type = parsed.type
    transaction.amount = parsed.amount
    transaction.currency = parsed.currency
    transaction.category = parsed.category
    transaction.merchant = parsed.merchant
    transaction.note = parsed.note
    transaction.occurred_on = parsed.occurred_on
    transaction.raw_text = raw_text
    await session.flush()
    return transaction


async def delete_transaction(session: AsyncSession, transaction: Transaction) -> None:
    await session.delete(transaction)
    await session.flush()


async def delete_all_user_transactions(session: AsyncSession, user: User) -> int:
    count = await session.scalar(select(func.count()).select_from(Transaction).where(Transaction.user_id == user.id))
    await session.execute(delete(Transaction).where(Transaction.user_id == user.id))
    await session.flush()
    return int(count or 0)


async def set_budget_limit(
    session: AsyncSession,
    user: User,
    category: str,
    amount: Decimal,
    currency: str,
) -> BudgetLimit:
    limit = await session.scalar(
        select(BudgetLimit).where(
            BudgetLimit.user_id == user.id,
            BudgetLimit.category == category,
            BudgetLimit.currency == currency,
        )
    )
    if limit:
        limit.amount = amount
    else:
        limit = BudgetLimit(user_id=user.id, category=category, amount=amount, currency=currency)
        session.add(limit)
    await session.flush()
    return limit


async def delete_budget_limit(session: AsyncSession, user: User, category: str, currency: str) -> None:
    await session.execute(
        delete(BudgetLimit).where(
            BudgetLimit.user_id == user.id,
            BudgetLimit.category == category,
            BudgetLimit.currency == currency,
        )
    )
    await session.flush()


async def list_budget_limits(session: AsyncSession, user: User) -> list[BudgetLimit]:
    rows = await session.scalars(
        select(BudgetLimit).where(BudgetLimit.user_id == user.id).order_by(BudgetLimit.category, BudgetLimit.currency)
    )
    return list(rows)


async def spending_by_category(
    session: AsyncSession,
    user: User,
    category: str,
    currency: str,
    from_date: date,
    to_date: date,
) -> Decimal:
    total = await session.scalar(
        select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            Transaction.user_id == user.id,
            Transaction.type == TransactionType.expense,
            Transaction.category == category,
            Transaction.currency == currency,
            Transaction.occurred_on >= from_date,
            Transaction.occurred_on <= to_date,
        )
    )
    return Decimal(total or 0)


async def totals_by_category(
    session: AsyncSession,
    user: User,
    from_date: date | None = None,
    to_date: date | None = None,
) -> list[tuple[str, TransactionType, str, Decimal]]:
    stmt = (
        select(Transaction.category, Transaction.type, Transaction.currency, func.sum(Transaction.amount))
        .where(Transaction.user_id == user.id)
        .group_by(Transaction.category, Transaction.type, Transaction.currency)
        .order_by(Transaction.type, Transaction.category, Transaction.currency)
    )
    stmt = _filter_dates(stmt, from_date, to_date)
    rows = await session.execute(stmt)
    return [(row[0], row[1], row[2], Decimal(row[3])) for row in rows.all()]


async def top_merchants(
    session: AsyncSession,
    user: User,
    from_date: date | None = None,
    to_date: date | None = None,
    limit: int = 5,
) -> list[tuple[str, str, Decimal]]:
    stmt = (
        select(Transaction.merchant, Transaction.currency, func.sum(Transaction.amount))
        .where(
            Transaction.user_id == user.id,
            Transaction.type == TransactionType.expense,
            Transaction.merchant.is_not(None),
        )
        .group_by(Transaction.merchant, Transaction.currency)
        .order_by(func.sum(Transaction.amount).desc())
        .limit(limit)
    )
    stmt = _filter_dates(stmt, from_date, to_date)
    rows = await session.execute(stmt)
    return [(row[0], row[1], Decimal(row[2])) for row in rows.all()]


async def total_by_merchant(session: AsyncSession, user: User, merchant: str) -> list[tuple[str, Decimal]]:
    rows = await session.execute(
        select(Transaction.currency, func.sum(Transaction.amount))
        .where(
            Transaction.user_id == user.id,
            func.lower(Transaction.merchant) == merchant.lower(),
        )
        .group_by(Transaction.currency)
        .order_by(Transaction.currency)
    )
    return [(row[0], Decimal(row[1])) for row in rows.all()]


async def export_transactions(
    session: AsyncSession,
    user: User,
    export_dir: Path,
    from_date: date | None = None,
    to_date: date | None = None,
) -> Path:
    export_dir.mkdir(parents=True, exist_ok=True)
    stmt = (
        select(Transaction)
        .where(Transaction.user_id == user.id)
        .order_by(Transaction.occurred_on.desc(), Transaction.id.desc())
    )
    stmt = _filter_dates(stmt, from_date, to_date)
    rows = await session.scalars(stmt)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(["Номер", "Дата", "Тип", "Сумма", "Валюта", "Категория", "Магазин/метка", "Заметка", "Исходный текст"])

    for tx in rows:
        ws.append(
            [
                tx.user_tx_number or tx.id,
                tx.occurred_on.isoformat(),
                "Доход" if tx.type == TransactionType.income else "Расход",
                float(tx.amount),
                tx.currency,
                category_label(tx.category),
                tx.merchant or "",
                tx.note or "",
                tx.raw_text,
            ]
        )

    summary_rows = await totals_by_category(session, user, from_date, to_date)
    summary = wb.create_sheet("Итоги по категориям")
    summary.append(["Тип", "Категория", "Валюта", "Итого"])
    for category, tx_type, currency, total in summary_rows:
        summary.append(
            [
                "Доход" if tx_type == TransactionType.income else "Расход",
                category_label(category),
                currency,
                float(total),
            ]
        )

    expense_rows = [row for row in summary_rows if row[1] == TransactionType.expense]
    chart_sheet = wb.create_sheet("График расходов")
    chart_sheet.append(["Категория", "Валюта", "Итого"])
    for category, _, currency, total in sorted(expense_rows, key=lambda row: row[3], reverse=True):
        chart_sheet.append([category_label(category), currency, float(total)])

    if len(expense_rows) > 0:
        chart = BarChart()
        chart.title = "Расходы по категориям"
        chart.y_axis.title = "Сумма"
        chart.x_axis.title = "Категория"
        data = Reference(chart_sheet, min_col=3, min_row=1, max_row=len(expense_rows) + 1)
        cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(expense_rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 18
        chart_sheet.add_chart(chart, "E2")

    merchant_rows = await top_merchants(
        session,
        user,
        from_date,
        to_date,
        limit=20,
    )
    merchants = wb.create_sheet("Топ магазинов")
    merchants.append(["Магазин/метка", "Валюта", "Итого"])
    for merchant, currency, total in merchant_rows:
        merchants.append([merchant, currency, float(total)])

    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 10), 42)

    suffix = ""
    if from_date and to_date:
        suffix = f"_{from_date.isoformat()}_{to_date.isoformat()}"
    path = export_dir / f"transactions_{user.telegram_id}{suffix}.xlsx"
    wb.save(path)
    return path


def _filter_dates(stmt: Select, from_date: date | None, to_date: date | None) -> Select:
    if from_date:
        stmt = stmt.where(Transaction.occurred_on >= from_date)
    if to_date:
        stmt = stmt.where(Transaction.occurred_on <= to_date)
    return stmt
